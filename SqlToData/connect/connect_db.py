#####################################################
# 実績データ変換プログラム
# 修正後はターミナルからpyinstaller connect_db.py --onefile
# を実行してexe化して配布すること。
#####################################################

import pymysql.cursors
import openpyxl as Excel

connection = pymysql.connect(host='192.168.3.203', user='kudo', password='1111',
                             db='sahashinewsystem', charset='utf8mb4',
                             cursorclass=pymysql.cursors.DictCursor)

try:
    def check_moji_len(moji):
        if len(moji) == 1:
            return '0' + str(moji)
        else:
            return str(moji)


    title = """
    ----------------------------------------------------------
    実績データ変更プログラム ver 0.0.2
    【変更点】
    金山　加硫実績の集計表示の変更依頼があったので別シートにて
    集計結果を表示するようにしました。
    （現時点では処理日から2ヶ月前までのデータが出力されます）
    ----------------------------------------------------------
    
    
    """
    print(title)

    print('化工品_加硫実績を出力します')
    wb = Excel.Workbook()
    ws = wb.active
    ws.title = '化工品_加硫実績'

    with connection.cursor() as cursor:
        sql = "SELECT * FROM kakouhin.df_karyujisseki"
        cursor.execute(sql)
        result = cursor.fetchall()
        ws['A1'] = 'ID'
        ws['B1'] = '加硫日'
        ws['C1'] = '直'
        ws['D1'] = '製番'
        ws['E1'] = 'ショット数'
        ws['F1'] = '型番'
        ws['G1'] = '不良コード1'
        ws['H1'] = '不良数1'
        ws['I1'] = '不良コード2'
        ws['J1'] = '不良数2'
        ws['K1'] = '不良コード3'
        ws['L1'] = '不良数3'
        ws['M1'] = '不良コード4'
        ws['N1'] = '不良数4'
        ws['O1'] = '停止コード1'
        ws['P1'] = '停止時間1'
        ws['Q1'] = '停止コード2'
        ws['R1'] = '停止時間2'
        ws['S1'] = '停止コード3'
        ws['T1'] = '停止時間3'
        ws['U1'] = '停止コード4'
        ws['V1'] = '停止時間4'
        ws['W1'] = 'ゴム配合1'
        ws['X1'] = 'ゴム配合2'
        ws['Y1'] = 'ゴム配合3'
        ws['Z1'] = 'ゴム配合4'
        ws['AA1'] = '処理日'
        i = 2
        for row in result:
            stri = str(i)
            ws['A' + stri] = row['ID']
            ws['B' + stri] = row['KARYUBI']
            ws['C' + stri] = row['TYOKU']
            ws['D' + stri] = row['SEIBAN']
            ws['E' + stri] = row['KARYU_SHOT']
            ws['F' + stri] = row['KATABAN']
            ws['G' + stri] = row['HURYO_CD1']
            ws['H' + stri] = row['HURYO_SU1']
            ws['I' + stri] = row['HURYO_CD2']
            ws['J' + stri] = row['HURYO_SU2']
            ws['K' + stri] = row['HURYO_CD3']
            ws['L' + stri] = row['HURYO_SU3']
            ws['M' + stri] = row['HURYO_CD4']
            ws['N' + stri] = row['HURYO_SU4']
            ws['O' + stri] = row['TEISI_CD1']
            ws['P' + stri] = row['TEISI_JIKAN1']
            ws['Q' + stri] = row['TEISI_CD2']
            ws['R' + stri] = row['TEISI_JIKAN2']
            ws['S' + stri] = row['TEISI_CD3']
            ws['T' + stri] = row['TEISI_JIKAN3']
            ws['U' + stri] = row['TEISI_CD4']
            ws['V' + stri] = row['TEISI_JIKAN4']
            ws['W' + stri] = row['GM1']
            ws['X' + stri] = row['GM2']
            ws['Y' + stri] = row['GM3']
            ws['Z' + stri] = row['GM4']
            ws['AA' + stri] = row['SYORIBI']
            i += 1
    print(str(i - 1) + '件 データ出力しました')

    print('検査実績を出力します')
    ws2 = wb.create_sheet(title='検査実績')
    with connection.cursor() as cursor:
        sql = "SELECT * FROM kakouhin.df_kakou_kensa"
        cursor.execute(sql)
        result = cursor.fetchall()
        ws2['A1'] = 'ID'
        ws2['B1'] = '検査部署'
        ws2['C1'] = '作業日'
        ws2['D1'] = '製番'
        ws2['E1'] = '仕入区分'
        ws2['F1'] = '個数'
        ws2['G1'] = '手直し'
        ws2['H1'] = '不良コード1'
        ws2['I1'] = '不良数1'
        ws2['J1'] = '不良コード2'
        ws2['K1'] = '不良数2'
        ws2['L1'] = '不良コード3'
        ws2['M1'] = '不良数3'
        ws2['N1'] = '不良コード4'
        ws2['O1'] = '不良数4'
        ws2['P1'] = '作業開始時間'
        ws2['Q1'] = '作業終了時間'
        ws2['R1'] = '入力日'
        i = 2
        for row in result:
            stri = str(i)
            ws2['A' + stri] = row['ID']
            ws2['B' + stri] = row['KENSABUSYO']
            ws2['C' + stri] = row['SAGYOUBI']
            ws2['D' + stri] = row['SEIBAN']
            ws2['E' + stri] = row['SIIREKUBUN']
            ws2['F' + stri] = row['SAGYOU_KOSU']
            ws2['G' + stri] = row['TENAOSI']
            ws2['H' + stri] = row['HURYO_CD1']
            ws2['I' + stri] = row['HURYO_SU1']
            ws2['J' + stri] = row['HURYO_CD2']
            ws2['K' + stri] = row['HURYO_SU2']
            ws2['L' + stri] = row['HURYO_CD3']
            ws2['M' + stri] = row['HURYO_SU3']
            ws2['N' + stri] = row['HURYO_CD4']
            ws2['O' + stri] = row['HURYO_SU4']
            if len(str(row['S_H_TIME'])) == 1:
                row['S_H_TIME'] = '0' + str(row['S_H_TIME'])
            else:
                row['S_H_TIME'] = str(row['S_H_TIME'])

            if len(str(row['S_M_TIME'])) == 1:
                row['S_M_TIME'] = '0' + str(row['S_M_TIME'])
            else:
                row['S_M_TIME'] = str(row['S_M_TIME'])

            if len(str(row['E_H_TIME'])) == 1:
                row['E_H_TIME'] = '0' + str(row['E_H_TIME'])
            else:
                row['E_H_TIME'] = str(row['E_H_TIME'])

            if len(str(row['E_M_TIME'])) == 1:
                row['E_M_TIME'] = '0' + str(row['E_M_TIME'])
            else:
                row['E_M_TIME'] = str(row['E_M_TIME'])

            ws2['P' + stri] = str(row['S_H_TIME']) + ':' + str(row['S_M_TIME'])
            ws2['Q' + stri] = str(row['E_H_TIME']) + ':' + str(row['E_M_TIME'])
            ws2['R' + stri] = row['INPUT']
            i += 1
    print(str(i - 1) + '件 データ出力しました')

    print('防振_加硫実績を出力します')
    ws3 = wb.create_sheet(title='防振_加硫実績')
    with connection.cursor() as cursor:
        sql = "SELECT * FROM karyu_keikaku.df_karyujisseki"
        cursor.execute(sql)
        result = cursor.fetchall()
        ws3['A1'] = 'ID'
        ws3['B1'] = '加硫日'
        ws3['C1'] = '加硫部署'
        ws3['D1'] = '原価コード'
        ws3['E1'] = '持ち台数'
        ws3['F1'] = '設備コード'
        ws3['G1'] = '号機コード'
        ws3['H1'] = '製番'
        ws3['I1'] = '型番'
        ws3['J1'] = '取数'
        ws3['K1'] = 'ショット数1'
        ws3['L1'] = 'ショット数2'
        ws3['M1'] = 'ショット数3'
        ws3['N1'] = '生産数'
        ws3['O1'] = '工程1'
        ws3['P1'] = '工程2'
        ws3['Q1'] = '工程3'
        ws3['R1'] = '不良コード1'
        ws3['S1'] = '不良数1'
        ws3['T1'] = '不良コード2'
        ws3['U1'] = '不良数2'
        ws3['V1'] = '不良コード3'
        ws3['W1'] = '不良数3'
        ws3['X1'] = '不良コード4'
        ws3['Y1'] = '不良数4'
        ws3['Z1'] = '不良コード5'
        ws3['AA1'] = '不良数5'
        ws3['AB1'] = '稼働時間'
        ws3['AC1'] = '稼働時間（分）'
        ws3['AD1'] = '停止コード1'
        ws3['AE1'] = '停止時間'
        ws3['AF1'] = '停止時間（分）'
        ws3['AG1'] = '停止コード2'
        ws3['AH1'] = '停止時間'
        ws3['AI1'] = '停止時間（分）'
        ws3['AJ1'] = '停止コード3'
        ws3['AK1'] = '停止時間'
        ws3['AL1'] = '停止時間（分）'
        ws3['AM1'] = '停止コード4'
        ws3['AN1'] = '停止時間'
        ws3['AO1'] = '停止時間（分）'
        ws3['AP1'] = '処理日'
        i = 2
        for row in result:
            stri = str(i)
            ws3['A' + stri] = row['ID']
            ws3['B' + stri] = row['KARYU_BI']
            ws3['C' + stri] = row['KARYU_BUSYO']
            ws3['D' + stri] = row['GENKA_CODE']
            ws3['E' + stri] = row['MOTIDAISU']
            ws3['F' + stri] = row['SETUBI_CODE']
            ws3['G' + stri] = row['GOUKI_CODE']
            ws3['H' + stri] = row['SEIBAN']
            ws3['I' + stri] = str(row['KATABAN_S']) + '～' + str(row['KATABAN_E'])
            ws3['J' + stri] = row['TORISU']
            ws3['K' + stri] = row['SHOT1']
            ws3['L' + stri] = row['SHOT2']
            ws3['M' + stri] = row['SHOT3']
            ws3['N' + stri] = row['SEISAN_SU']
            ws3['O' + stri] = row['KOUTEI1']
            ws3['P' + stri] = row['KOUTEI2']
            ws3['Q' + stri] = row['KOUTEI3']
            ws3['R' + stri] = row['HURYO_CODE1']
            ws3['S' + stri] = row['HURYO_SU1']
            ws3['T' + stri] = row['HURYO_CODE2']
            ws3['U' + stri] = row['HURYO_SU2']
            ws3['V' + stri] = row['HURYO_CODE3']
            ws3['W' + stri] = row['HURYO_SU3']
            ws3['X' + stri] = row['HURYO_CODE4']
            ws3['Y' + stri] = row['HURYO_SU4']
            ws3['Z' + stri] = row['HURYO_CODE5']
            ws3['AA' + stri] = row['HURYO_SU5']

            row['KSH'] = check_moji_len(str(row['KSH']))
            row['KSM'] = check_moji_len(str(row['KSM']))
            row['KEH'] = check_moji_len(str(row['KEH']))
            row['KEM'] = check_moji_len(str(row['KEM']))

            ws3['AB' + stri] = str(row['KSH']) + ':' + str(row['KSM']) + '～' + str(row['KEH']) + ':' + str(row['KEM'])
            ws3['AC' + stri] = row['KADOUJIKAN']
            ws3['AD' + stri] = row['TEISI_CD1']

            row['TSH1'] = check_moji_len(str(row['TSH1']))
            row['TSM1'] = check_moji_len(str(row['TSM1']))
            row['TEH1'] = check_moji_len(str(row['TEH1']))
            row['TEM1'] = check_moji_len(str(row['TEM1']))

            ws3['AE' + stri] = str(row['TSH1']) + ':' + str(row['TSM1']) + '～' + str(row['TEH1']) + ':' + str(
                row['TEM1'])
            ws3['AF' + stri] = row['TEISI_JIKAN1']
            ws3['AG' + stri] = row['TEISI_CD2']

            row['TSH2'] = check_moji_len(str(row['TSH2']))
            row['TSM2'] = check_moji_len(str(row['TSM2']))
            row['TEH2'] = check_moji_len(str(row['TEH2']))
            row['TEM2'] = check_moji_len(str(row['TEM2']))

            ws3['AH' + stri] = str(row['TSH2']) + ':' + str(row['TSM2']) + '～' + str(row['TEH2']) + ':' + str(
                row['TEM2'])
            ws3['AI' + stri] = row['TEISI_JIKAN2']
            ws3['AJ' + stri] = row['TEISI_CD3']

            row['TSH3'] = check_moji_len(str(row['TSH3']))
            row['TSM3'] = check_moji_len(str(row['TSM3']))
            row['TEH3'] = check_moji_len(str(row['TEH3']))
            row['TEM3'] = check_moji_len(str(row['TEM3']))

            ws3['AK' + stri] = str(row['TSH3']) + ':' + str(row['TSM3']) + '～' + str(row['TEH3']) + ':' + str(
                row['TEM3'])
            ws3['AL' + stri] = row['TEISI_JIKAN3']
            ws3['AM' + stri] = row['TEISI_CD4']

            row['TSH4'] = check_moji_len(str(row['TSH4']))
            row['TSM4'] = check_moji_len(str(row['TSM4']))
            row['TEH4'] = check_moji_len(str(row['TEH4']))
            row['TEM4'] = check_moji_len(str(row['TEM4']))

            ws3['AN' + stri] = str(row['TSH4']) + ':' + str(row['TSM4']) + '～' + str(row['TEH4']) + ':' + str(
                row['TEM4'])
            ws3['AO' + stri] = row['TEISI_JIKAN4']
            ws3['AP' + stri] = row['SYORIBI']
            i += 1
    print(str(i - 1) + '件 データ出力しました')

    print('仕入_不良を出力します')
    ws4 = wb.create_sheet(title='仕入_不良')
    with connection.cursor() as cursor:
        sql = "SELECT * FROM sahashinewsystem.df_siirejisseki_kensa"
        cursor.execute(sql)
        result = cursor.fetchall()
        ws4['A1'] = 'ID'
        ws4['B1'] = '検査日'
        ws4['C1'] = '納入者'
        ws4['D1'] = '納品日'
        ws4['E1'] = '伝票番号'
        ws4['F1'] = '納入先'
        ws4['G1'] = '製番'
        ws4['H1'] = '仕入区分'
        ws4['I1'] = '数量'
        ws4['J1'] = '箱数'
        ws4['K1'] = '不良数'
        ws4['L1'] = '不良コード1'
        ws4['M1'] = '不良数1'
        ws4['N1'] = '不良コード2'
        ws4['O1'] = '不良数2'
        ws4['P1'] = '不良コード3'
        ws4['Q1'] = '不良数3'
        ws4['R1'] = '不良コード4'
        ws4['S1'] = '不良数4'
        ws4['T1'] = '不良コード5'
        ws4['U1'] = '不良数5'
        ws4['V1'] = '合格数'
        ws4['W1'] = '処理日'
        ws4['X1'] = '有効'
        i = 2
        for row in result:
            stri = str(i)
            ws4['A' + stri] = row['SIIRE_ID']
            ws4['B' + stri] = str(row['KENSA_BI'])[0:10]
            ws4['C' + stri] = row['NOUNYUSYA']
            ws4['D' + stri] = str(row['NOUHIN_BI'])[0:10]
            ws4['E' + stri] = row['DENPYOU_NUMBER']
            ws4['F' + stri] = row['NOUNYUSAKI']
            ws4['G' + stri] = row['SEIBAN']
            ws4['H' + stri] = row['SIIREKUBUN']
            ws4['I' + stri] = row['SUURYOU']
            ws4['J' + stri] = row['HAKOSU']
            ws4['K' + stri] = row['HURYOU_SU']
            ws4['L' + stri] = row['HURYO_CD1']
            ws4['M' + stri] = row['HURYO_SU1']
            ws4['N' + stri] = row['HURYO_CD2']
            ws4['O' + stri] = row['HURYO_SU2']
            ws4['P' + stri] = row['HURYO_CD3']
            ws4['Q' + stri] = row['HURYO_SU3']
            ws4['R' + stri] = row['HURYO_CD4']
            ws4['S' + stri] = row['HURYO_SU4']
            ws4['T' + stri] = row['HURYO_CD5']
            ws4['U' + stri] = row['HURYO_SU5']
            ws4['V' + stri] = row['GOUKAKU_SU']
            ws4['W' + stri] = str(row['SYORIBI'])[0:10]
            ws4['X' + stri] = row['YUUKOU']
            i += 1
    print(str(i - 1) + '件 データ出力しました')

    print('金山防振_加硫実績を出力します')
    ws4 = wb.create_sheet(title='金山防振_加硫実績')
    with connection.cursor() as cursor:
        sql = "SELECT *,mns.SEITNK " \
              "FROM karyu_keikaku.df_karyujisseki dk " \
              "LEFT JOIN sahashinewsystem.mf_new_seihintanka mns on dk.SEIBAN=mns.SEIBAN " \
              "WHERE KARYU_BUSYO=2000 and KARYU_BI >=now()-interval 2 month " \
              "order by dk.KARYU_BI,dk.SEIBAN asc,dk.SETUBI_CODE asc,dk.GOUKI_CODE asc"
        cursor.execute(sql)
        result = cursor.fetchall()
        ws4['A1'] = '製番'
        ws4['B1'] = '設備コード'
        ws4['C1'] = '号機コード'
        ws4['D1'] = '加硫日'
        ws4['E1'] = '不良コード'
        ws4['F1'] = '不良名'
        ws4['G1'] = '不良数'
        ws4['H1'] = '生産数'
        ws4['I1'] = '単価'
        ws4['J1'] = '金額'
        i = 2
        for row in result:
            stri = str(i)
            ws4['A' + stri] = row['SEIBAN']
            ws4['B' + stri] = row['SETUBI_CODE']
            ws4['C' + stri] = row['GOUKI_CODE']
            ws4['D' + stri] = row['KARYU_BI']
            ws4['E' + stri] = row['HURYO_CODE1']
            sql2 = "SELECT HURYO_NAME FROM sahashinewsystem.mf_huryocode WHERE CD = %s"
            cursor.execute(sql2, row['HURYO_CODE1'])
            result2 = cursor.fetchone()
            if result2 != None:
                ws4['F' + stri] = result2['HURYO_NAME']
            else:
                ws4['F' + stri] = '-'
            ws4['G' + stri] = row['HURYO_SU1']
            ws4['H' + stri] = row['SEISAN_SU']
            ws4['I' + stri] = row['SEITNK']
            ws4['J' + stri] = round(row['HURYO_SU1'] * row['SEITNK'])
            i += 1
            for j in range(2, 5):
                stri = str(i)
                strj = str(j)
                if row['HURYO_CODE' + strj] != 0:
                    ws4['A' + stri] = row['SEIBAN']
                    ws4['B' + stri] = row['SETUBI_CODE']
                    ws4['C' + stri] = row['GOUKI_CODE']
                    ws4['D' + stri] = row['KARYU_BI']
                    ws4['E' + stri] = row['HURYO_CODE' + strj]
                    sql2 = "SELECT HURYO_NAME FROM sahashinewsystem.mf_huryocode WHERE CD = %s"
                    key = row['HURYO_CODE' + strj]
                    cursor.execute(sql2, key)
                    result2 = cursor.fetchone()
                    if result2 != None:
                        ws4['F' + stri] = result2['HURYO_NAME']
                    else:
                        ws4['F' + stri] = '-'
                    ws4['G' + stri] = row['HURYO_SU' + strj]
                    ws4['H' + stri] = 0
                    ws4['I' + stri] = row['SEITNK']
                    ws4['J' + stri] = round(row['HURYO_SU' + strj] * row['SEITNK'])
                    i += 1
                    j += j

    print(str(i - 1) + '件 データ出力しました')

    wb.save('karyu_kensa_jisseki.xlsx')
    connection.commit()

finally:
    connection.close()
