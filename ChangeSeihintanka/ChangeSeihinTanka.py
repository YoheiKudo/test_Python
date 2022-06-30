#####################################################
# 実績データ変換プログラム
# 修正後はターミナルからpyinstaller ChangeSeihinTanka.py --onefile
# を実行してexe化して配布すること。
#####################################################

import pymysql.cursors
import openpyxl as Excel
from tqdm import tqdm

title = """
----------------------------------------------------------
製品単価マスター変換システム ver 1.0.0
----------------------------------------------------------
"""

print(title)

password = input('パスワードを入力してください：　')

if password == '00709200' or password == '01022960' or password == '00331950':
    def changesetubicode(sc):
        if sc == '01':
            return 'MTI'
        elif sc == '11':
            return 'SSI'
        elif sc == '21':
            return 'AIﾀﾃ'
        elif sc == '31':
            return 'AIﾖｺ'
        elif sc == '51':
            return 'STI'
        elif sc == '60':
            return '特殊'
        elif sc == 'LTI':
            return 'STI'
        else:
            return sc


    connection = pymysql.connect(host='192.168.3.203', user='kudo', password='1111',
                                 db='sahashinewsystem', charset='utf8mb4',
                                 cursorclass=pymysql.cursors.DictCursor)

    try:
        def check_moji_len(moji):
            if len(moji) == 1:
                return '0' + str(moji)
            else:
                return str(moji)


        caution_message = """
        製品単価マスター.xlsxファイルが開いているとエラーになります。
        必ず閉じてからください
        """
        print(caution_message)

        print('製品単価マスターを出力します')
        wb = Excel.Workbook()
        ws = wb.active
        ws.title = '化工品_加硫実績'

        with connection.cursor() as cursor:
            sql = "SELECT * FROM sahashinewsystem.mf_new_seihintanka"
            cursor.execute(sql)
            result = cursor.fetchall()
            ws['A1'] = '製番'
            ws['B1'] = '単価区分'
            ws['C1'] = '正単価'
            ws['D1'] = '仮単価'
            ws['E1'] = '車種'
            ws['F1'] = '品種'
            ws['G1'] = 'ゴム配合1'
            ws['H1'] = 'ゴムスリット1'
            ws['I1'] = 'ゴム重量1'
            ws['J1'] = 'ゴム単価1'
            ws['K1'] = 'ゴム材料単価1'
            ws['L1'] = 'ゴム単価区分1'
            ws['M1'] = 'ゴム単価重量1'
            ws['N1'] = 'ゴム配合2'
            ws['O1'] = 'ゴムスリット2'
            ws['P1'] = 'ゴム重量2'
            ws['Q1'] = 'ゴム単価2'
            ws['R1'] = 'ゴム材料単価2'
            ws['S1'] = 'ゴム単価区分2'
            ws['T1'] = 'ゴム単価重量2'
            ws['U1'] = 'ゴム材料計'
            ws['V1'] = '金具製番1'
            ws['W1'] = '金具単価1'
            ws['X1'] = '金具単価区分1'
            ws['Y1'] = '金具仕入単価1'
            ws['Z1'] = '金具個数1'
            ws['AA1'] = '金具金額1'
            ws['AB1'] = '金具製番2'
            ws['AC1'] = '金具単価2'
            ws['AD1'] = '金具単価区分2'
            ws['AE1'] = '金具仕入単価2'
            ws['AF1'] = '金具個数2'
            ws['AG1'] = '金具金額2'
            ws['AH1'] = '金具製番3'
            ws['AI1'] = '金具単価3'
            ws['AJ1'] = '金具単価区分3'
            ws['AK1'] = '金具仕入単価3'
            ws['AL1'] = '金具個数3'
            ws['AM1'] = '金具金額3'
            ws['AN1'] = '金具製番4'
            ws['AO1'] = '金具単価4'
            ws['AP1'] = '金具単価区分4'
            ws['AQ1'] = '金具仕入単価4'
            ws['AR1'] = '金具個数4'
            ws['AS1'] = '金具金額4'
            ws['AT1'] = '金具製番5'
            ws['AU1'] = '金具単価5'
            ws['AV1'] = '金具単価区分5'
            ws['AW1'] = '金具仕入単価5'
            ws['AX1'] = '金具個数5'
            ws['AY1'] = '金具金額5'
            ws['AZ1'] = '接着剤'
            ws['BA1'] = '金具材料計'
            ws['BB1'] = '総材料計'
            ws['BC1'] = '材料比率'
            ws['BD1'] = '化成被膜'
            ws['BE1'] = '金具成型'
            ws['BF1'] = 'ゴム成型'
            ws['BG1'] = '加硫'
            ws['BH1'] = '機種'
            ws['BI1'] = '取数'
            ws['BJ1'] = '時間(分)'
            ws['BK1'] = '時間(秒)'
            ws['BL1'] = '成型計'
            ws['BM1'] = '出来高'
            ws['BN1'] = 'バリ'
            ws['BO1'] = '塗装'
            ws['BP1'] = '絞り'
            ws['BQ1'] = 'シリコン'
            ws['BR1'] = '圧入'
            ws['BS1'] = '識別'
            ws['BT1'] = 'ASSY'
            ws['BU1'] = 'ノックスラスト'
            ws['BV1'] = 'キャップ抜き'
            ws['BW1'] = '切り割り'
            ws['BX1'] = 'エフ付け'
            ws['BY1'] = '梱包'
            ws['BZ1'] = 'その他1名称'
            ws['CA1'] = 'その他1単価'
            ws['CB1'] = 'その他2名称'
            ws['CC1'] = 'その他2単価'
            ws['CD1'] = 'その他3名称'
            ws['CE1'] = 'その他3単価'
            ws['CF1'] = '検査'
            ws['CG1'] = '仕上計'
            ws['CH1'] = '加工計'
            ws['CI1'] = '加工比'
            ws['CJ1'] = '直接原価'
            ws['CK1'] = '直原比'
            ws['CL1'] = '償却'
            ws['CM1'] = '間接費'
            ws['CN1'] = '一般管理'
            ws['CO1'] = '利益'
            ws['CP1'] = '利益率'
            ws['CQ1'] = '製品単価'
            ws['CR1'] = '仕掛単価'
            ws['CS1'] = 'チェック'
            ws['CT1'] = '加硫時間（分）'
            ws['CU1'] = '加硫時間（秒）'
            ws['CV1'] = '旧製番'
            ws['CW1'] = '新製番'
            ws['CX1'] = '備考'
            ws['CY1'] = '製品重量'
            ws['CZ1'] = '金具重量1'
            ws['DA1'] = '金具重量2'
            ws['DB1'] = '金具重量3'
            ws['DC1'] = '処理日'
            ws['DD1'] = '品番'

            i = 2
            for row in tqdm(result):
                stri = str(i)
                ws['A' + stri] = row['SEIBAN']
                ws['B' + stri] = row['TNKKBN']
                ws['C' + stri] = row['SEITNK']
                ws['D' + stri] = row['KRITNK']
                ws['E' + stri] = row['SYASYU']
                ws['F' + stri] = row['HINSYU']
                ws['G' + stri] = row['GMHAIGOU1']
                ws['H' + stri] = row['GMSLIT1']
                ws['I' + stri] = row['GMJYURYO1']
                ws['J' + stri] = row['GMTNK1']
                ws['K' + stri] = row['GMZAITNK1']
                ws['L' + stri] = row['GMTNKKBN1']
                ws['M' + stri] = row['GMTNKJYURYO1']
                ws['N' + stri] = row['GMHAIGOU2']
                ws['O' + stri] = row['GMSLIT2']
                ws['P' + stri] = row['GMJYURYO2']
                ws['Q' + stri] = row['GMTNK2']
                ws['R' + stri] = row['GMZAITNK2']
                ws['S' + stri] = row['GMTNKKBN2']
                ws['T' + stri] = row['GMTNKJYURYO2']
                ws['U' + stri] = row['GMZAIKEI']
                ws['V' + stri] = row['KGSBN1']
                ws['W' + stri] = row['KGTNK1']
                ws['X' + stri] = row['KGTNKKBN1']
                ws['Y' + stri] = row['KGSIRTNK1']
                ws['Z' + stri] = row['KGKOSU1']
                ws['AA' + stri] = row['KGKNGK1']
                ws['AB' + stri] = row['KGSBN2']
                ws['AC' + stri] = row['KGTNK2']
                ws['AD' + stri] = row['KGTNKKBN2']
                ws['AE' + stri] = row['KGSIRTNK2']
                ws['AF' + stri] = row['KGKOSU2']
                ws['AG' + stri] = row['KGKNGK2']
                ws['AH' + stri] = row['KGSBN3']
                ws['AI' + stri] = row['KGTNK3']
                ws['AJ' + stri] = row['KGTNKKBN3']
                ws['AK' + stri] = row['KGSIRTNK3']
                ws['AL' + stri] = row['KGKOSU3']
                ws['AM' + stri] = row['KGKNGK3']
                ws['AN' + stri] = row['KGSBN4']
                ws['AO' + stri] = row['KGTNK4']
                ws['AP' + stri] = row['KGTNKKBN4']
                ws['AQ' + stri] = row['KGSIRTNK4']
                ws['AR' + stri] = row['KGKOSU4']
                ws['AS' + stri] = row['KGKNGK4']
                ws['AT' + stri] = row['KGSBN5']
                ws['AU' + stri] = row['KGTNK5']
                ws['AV' + stri] = row['KGTNKKBN5']
                ws['AW' + stri] = row['KGSIRTNK5']
                ws['AX' + stri] = row['KGKOSU5']
                ws['AY' + stri] = row['KGKNGK5']
                ws['AZ' + stri] = row['STKZAI']
                ws['BA' + stri] = row['KGZAIKEI']
                ws['BB' + stri] = row['TOTAL_ZAIKEI']
                ws['BC' + stri] = row['ZAIHI']
                ws['BD' + stri] = row['KSIHMKKEI']
                ws['BE' + stri] = row['KGSEIKEIKEI']
                ws['BF' + stri] = row['GMSEIKEI']
                ws['BG' + stri] = row['KARYU']
                ws['BH' + stri] = row['KISYU']
                ws['BI' + stri] = row['TORISU']
                ws['BJ' + stri] = row['MTIMEM']
                ws['BK' + stri] = row['MTIMES']
                ws['BL' + stri] = row['SEIKEIKEI']
                ws['BM' + stri] = row['DEKIDAKA']
                ws['BN' + stri] = row['BARI']
                ws['BO' + stri] = row['TOSOU']
                ws['BP' + stri] = row['SIBORI']
                ws['BQ' + stri] = row['SILICONE']
                ws['BR' + stri] = row['ATUNYU']
                ws['BS' + stri] = row['SIKIBETU']
                ws['BT' + stri] = row['ASSY']
                ws['BU' + stri] = row['NOXRUST']
                ws['BV' + stri] = row['CAP']
                ws['BW' + stri] = row['CUT']
                ws['BX' + stri] = row['TAG']
                ws['BY' + stri] = row['PACKING']
                ws['BZ' + stri] = row['OTHER1_NAME']
                ws['CA' + stri] = row['OTHER1']
                ws['CB' + stri] = row['OTHER2_NAME']
                ws['CC' + stri] = row['OTHER2']
                ws['CD' + stri] = row['OTHER3_NAME']
                ws['CE' + stri] = row['OTHER3']
                ws['CF' + stri] = row['KENSA']
                ws['CG' + stri] = row['SIAGEKEI']
                ws['CH' + stri] = row['KAKOUKEI']
                ws['CI' + stri] = row['KAKOUHI']
                ws['CJ' + stri] = row['TYOKUGEN']
                ws['CK' + stri] = row['TYOKUGENHI']
                ws['CL' + stri] = row['SYOKYAKU']
                ws['CM' + stri] = row['KANSETU']
                ws['CN' + stri] = row['IPPAN']
                ws['CO' + stri] = row['RIEKI']
                ws['CP' + stri] = row['RIEKIRITU']
                ws['CQ' + stri] = row['SHNTNK']
                ws['CR' + stri] = row['SIKATNK']
                ws['CS' + stri] = row['CEK']
                ws['CT' + stri] = row['KTIMEM']
                ws['CU' + stri] = row['KTIMES']
                ws['CV' + stri] = row['KYUSBN']
                ws['CW' + stri] = row['SINSBN']
                ws['CX' + stri] = row['BIKOU']
                ws['CY' + stri] = row['SEIHINJYURYO']
                ws['CZ' + stri] = row['KG1JYURYO']
                ws['DA' + stri] = row['KG2JYURYO']
                ws['DB' + stri] = row['KG3JYURYO']
                ws['DC' + stri] = row['SYORIBI']
                ws['DD' + stri] = row['HINBAN']
                i += 1
        print(str(i - 2) + '件 データ出力しました\n')

        wb.save('製品単価マスター.xlsx')
        connection.commit()

    except Exception:
        print('Error')
    finally:
        connection.close()
else:
    print('パスワードが違います')
